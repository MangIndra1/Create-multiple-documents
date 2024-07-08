import os
import shutil
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from docx2pdf import convert

# Menghubungkan dengan folder (Buat Folder di tempat yang sama dengan program)
Sp_folder = 'Sponsorship'
if not os.path.exists(Sp_folder):
    os.makedirs(Sp_folder)

# Menghubungkan sheet yang dipakai (Tulis Template excel yang digunakan)
wb = load_workbook(filename='Nama Perusahaan.xlsx', read_only=True, data_only=True)
sheet_name = 'Sheet1'
ws = wb[sheet_name]

# min_row = baris awal data yang mau diambil di excel (min_row = 2 , dimulai dari baris ke-2)
for row in ws.iter_rows(min_row=2, values_only=True):
    # row[x] = kolom yang ingin diambil + 1. row[2] berarti mengambil data di kolom ke-3
    nama_sponsorship = row[1]
    # Jika row kosong
    if nama_sponsorship:
        docx_template = "Template Surat Pengantar Permohonan Proposal Sponsorship 2024.docx"
        
        if not os.path.exists(docx_template):
            print(f"Template '{docx_template}' tidak ditemukan.")
            continue
        # File Sponsor (Sesuaikan context pada template Word yang ingin diedit)
        Sponsor_doc = DocxTemplate(docx_template)
        Sponsor_context = {'nama_sponsor': nama_sponsorship}
        Sponsor_doc.render(Sponsor_context)
        
        # ubah nama file (Sesuaikan perubahan nama file sesuai kebutuhan)
        Sponsor_file_name = f"6 - {nama_sponsorship}" 
        
        # simpan file ke folder yang dituju
        Sponsor_docx_path = Sponsor_file_name + ".docx"
        Sponsor_doc.save(Sponsor_docx_path)
        
        # convert Word ke pdf
        convert(Sponsor_docx_path)
        Sponsor_pdf_path = Sponsor_file_name + ".pdf"
        new_Sponsor_pdf_path = os.path.join(Sp_folder, Sponsor_pdf_path)
        shutil.move(Sponsor_pdf_path, new_Sponsor_pdf_path)